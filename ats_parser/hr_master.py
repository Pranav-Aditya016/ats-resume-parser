"""The recruiter-facing workbook: an HR Candidate Master sheet plus a Raw Data sheet.

Recruiters shortlist from a narrow, consistent set of columns; auditors and analysts
need everything. Splitting those into two sheets lets each be right for its reader
instead of compromising on one 155-column sheet nobody can scan.

Employment history is capped at the most recent three positions. A candidate with
twenty roles produced twenty column groups before, which made the sheet unusable
horizontally while adding nothing to a shortlisting decision. The full history stays
on the Raw Data sheet, and Total Positions Listed says how many were truncated.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ats_parser.experience import summarize_experience

MAX_POSITIONS = 3
NOT_IN_RESUME = "Not mentioned in resume"

# Ordered most to least advanced, so the first match is the highest qualification.
QUALIFICATION_ORDER = [
    ("Doctorate", ("phd", "ph.d", "doctor")),
    ("Post Graduate", ("master", "m.tech", "mtech", "mba", "mca", "msc", "m.sc", "m.com", "m.e", "m.arch")),
    ("Under Graduate", ("bachelor", "b.tech", "btech", "b.e", "be ", "bsc", "b.sc", "b.com", "bba", "bca", "llb", "mbbs", "b.arch")),
    ("Diploma", ("diploma",)),
    ("Higher Secondary", ("higher secondary", "12th", "hsc", "puc", "intermediate")),
    ("Secondary", ("secondary", "sslc", "10th")),
]

MASTER_COLUMNS = [
    "Resume ID", "Source File", "Full Name", "Email", "Phone",
    "Current Location", "Preferred Location",
    "Total Experience (Years)", "Experience Basis",
    "Current Designation", "Current Company",
    "Key Skills", "Software / Tools", "Soft Skills", "Languages",
    "Highest Qualification", "Degree", "Specialization", "Institution", "Graduation Year",
]
for _position in range(1, MAX_POSITIONS + 1):
    MASTER_COLUMNS += [
        f"Job {_position} Company", f"Job {_position} Designation",
        f"Job {_position} From", f"Job {_position} To", f"Job {_position} Duration",
    ]
MASTER_COLUMNS += [
    "Total Positions Listed", "Certifications", "Professional Summary",
    # Filled by the recruiter after speaking to the candidate, as the change
    # request asks: show the gap rather than leaving the cell silently blank.
    "Current CTC", "Expected CTC", "Notice Period",
    "Applied Position", "Resume Source", "Screening Status", "Recruiter Remarks",
]

SKIP_FILES = {"batch_metrics.json", "failures.json"}


def load_records(input_dir: Path) -> list[tuple[str, dict]]:
    records = []
    for path in sorted(input_dir.glob("*.json")):
        if path.name in SKIP_FILES or path.name.endswith(".usage.json"):
            continue
        try:
            record = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if isinstance(record, dict):
            records.append((path.stem, record))
    return records


def master_row(index: int, source: str, record: dict) -> dict:
    personal = record.get("personal_information") or {}
    skills = record.get("skills") or {}
    positions = ordered_positions(record.get("work_experience") or [])
    summary = record.get("experience_summary") or summarize_experience(record.get("work_experience") or [])
    education = highest_education(record.get("education") or [])

    row = {
        "Resume ID": index,
        "Source File": source,
        "Full Name": personal.get("full_name"),
        "Email": personal.get("email"),
        "Phone": personal.get("phone"),
        "Current Location": personal.get("location"),
        "Preferred Location": NOT_IN_RESUME,
        "Total Experience (Years)": summary.get("total_experience_years"),
        "Experience Basis": summary.get("computation_basis"),
        "Current Designation": summary.get("current_designation"),
        "Current Company": summary.get("current_company"),
        "Key Skills": _join(_key_skills(skills)),
        "Software / Tools": _join(skills.get("developer_tools")),
        "Soft Skills": _join(skills.get("soft_skills")),
        "Languages": _join(record.get("languages")),
        "Highest Qualification": education["qualification"],
        "Degree": education["degree"],
        "Specialization": education["specialization"],
        "Institution": education["institution"],
        "Graduation Year": education["year"],
        "Total Positions Listed": len(positions) or None,
        "Certifications": _join(c.get("name") for c in record.get("certifications") or [] if isinstance(c, dict)),
        "Professional Summary": record.get("professional_summary"),
        "Current CTC": NOT_IN_RESUME,
        "Expected CTC": NOT_IN_RESUME,
        "Notice Period": NOT_IN_RESUME,
    }
    for slot in range(MAX_POSITIONS):
        entry = positions[slot] if slot < len(positions) else {}
        number = slot + 1
        row[f"Job {number} Company"] = entry.get("company")
        row[f"Job {number} Designation"] = entry.get("role")
        row[f"Job {number} From"] = entry.get("start_date")
        row[f"Job {number} To"] = entry.get("end_date")
        row[f"Job {number} Duration"] = entry.get("duration")
    return {column: row.get(column) for column in MASTER_COLUMNS}


def ordered_positions(entries: list[dict]) -> list[dict]:
    """Most recent first, so slot 1 is always the latest role."""
    return sorted((e for e in entries if isinstance(e, dict)), key=_recency, reverse=True)


def _recency(entry: dict) -> tuple[int, int]:
    end = _sortable(entry.get("end_date"))
    start = _sortable(entry.get("start_date"))
    return (end, start)


def _sortable(value: Any) -> int:
    if not isinstance(value, str) or not value.strip():
        return 0
    text = value.strip()
    if text.lower() == "present":
        return 999_999  # An ongoing role always sorts newest.
    digits = "".join(ch for ch in text[:7] if ch.isdigit())
    return int(digits.ljust(6, "0")[:6]) if digits else 0


def highest_education(entries: list[dict]) -> dict:
    for label, markers in QUALIFICATION_ORDER:
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            degree = str(entry.get("degree") or "").lower()
            if any(marker in degree for marker in markers):
                return {
                    "qualification": label,
                    "degree": entry.get("degree"),
                    "specialization": entry.get("specialization"),
                    "institution": entry.get("institution"),
                    "year": _graduation_year(entry),
                }
    return {"qualification": None, "degree": None, "specialization": None, "institution": None, "year": None}


def _graduation_year(entry: dict) -> str | None:
    for field in ("end_date", "start_date"):
        value = str(entry.get(field) or "")
        if len(value) >= 4 and value[:4].isdigit():
            return value[:4]
    return None


def _key_skills(skills: dict) -> list:
    """Everything a recruiter would scan, minus tools and soft skills."""
    collected = []
    for key in ("other_skills", "programming_languages", "frameworks", "libraries",
                "databases", "cloud_platforms"):
        collected.extend(skills.get(key) or [])
    seen, unique = set(), []
    for value in collected:
        marker = str(value).casefold()
        if value and marker not in seen:
            unique.append(value)
            seen.add(marker)
    return unique


def raw_row(index: int, source: str, record: dict) -> dict:
    """Every extracted value, flattened. The audit trail behind the Master sheet."""
    personal = record.get("personal_information") or {}
    skills = record.get("skills") or {}
    stats = record.get("resume_statistics") or {}
    row = {
        "Resume ID": index,
        "Source File": source,
        **{f"Personal: {k}": v for k, v in personal.items()},
        "Professional Summary": record.get("professional_summary"),
        **{f"Skills: {k}": _join(v) for k, v in skills.items()},
    }
    for field in ("languages", "achievements", "leadership", "publications", "patents",
                  "hackathons", "competitions", "awards", "volunteer_experience",
                  "extracurricular_activities", "references", "keywords", "ats_score_keywords"):
        row[field.replace("_", " ").title()] = _join(record.get(field))

    for slot, entry in enumerate(record.get("education") or [], start=1):
        if isinstance(entry, dict):
            for key, value in entry.items():
                row[f"Education {slot}: {key}"] = value

    for slot, entry in enumerate(ordered_positions(record.get("work_experience") or []), start=1):
        for key, value in entry.items():
            row[f"Experience {slot}: {key}"] = _join(value) if isinstance(value, list) else value

    for label, field in (("Internship", "internships"), ("Project", "projects"), ("Certification", "certifications")):
        for slot, entry in enumerate(record.get(field) or [], start=1):
            if isinstance(entry, dict):
                for key, value in entry.items():
                    row[f"{label} {slot}: {key}"] = _join(value) if isinstance(value, list) else value

    summary = record.get("experience_summary") or summarize_experience(record.get("work_experience") or [])
    row.update({f"Derived: {k}": v for k, v in summary.items()})
    row.update({f"Stats: {k}": _join(v) if isinstance(v, list) else v for k, v in stats.items()})
    return row


def build_workbook(records: list[tuple[str, dict]]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    master = [master_row(i, source, record) for i, (source, record) in enumerate(records, start=1)]
    raw = [raw_row(i, source, record) for i, (source, record) in enumerate(records, start=1)]
    _add_sheet(workbook, "HR Candidate Master", master, MASTER_COLUMNS)
    _add_sheet(workbook, "Raw Data", raw, _union_columns(raw))
    return workbook


def _union_columns(rows: list[dict]) -> list[str]:
    """Preserve first-seen order across rows with differing shapes."""
    columns, seen = [], set()
    for row in rows:
        for key in row:
            if key not in seen:
                columns.append(key)
                seen.add(key)
    return columns


def _add_sheet(workbook: Workbook, title: str, rows: list[dict], columns: list[str]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([_cell(row.get(column)) for column in columns])
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for position, column in enumerate(columns, start=1):
        widest = max((len(str(row.get(column) or "")) for row in rows), default=0)
        sheet.column_dimensions[get_column_letter(position)].width = min(45, max(12, widest + 2, len(column) + 2))


def _cell(value: Any):
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _join(values: Any) -> str | None:
    if values is None:
        return None
    if isinstance(values, str):
        return values
    rendered = [str(v) for v in values if v is not None]
    return ", ".join(rendered) if rendered else None
